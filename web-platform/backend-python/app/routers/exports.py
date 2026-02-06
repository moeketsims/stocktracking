from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import Response
from typing import Optional
from datetime import datetime, timedelta
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF

from ..config import get_supabase_admin_client
from ..routers.auth import require_manager, get_view_location_id

router = APIRouter(prefix="/exports", tags=["Exports"])

# Shared styling constants
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="E5E7EB"),
)


def _style_header_row(ws):
    """Apply standard header styling to the first row."""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT


def _auto_column_widths(ws):
    """Auto-adjust column widths based on content."""
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 50)


def _workbook_to_response(wb, filename: str) -> Response:
    """Convert workbook to FastAPI Response."""
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------- Stock Balance Export ----------

@router.get("/stock/excel")
async def export_stock_excel(
    view_location_id: Optional[str] = None,
    user_data: dict = Depends(require_manager),
):
    """Export consolidated stock balance across all locations as Excel."""
    supabase = get_supabase_admin_client()

    # Fetch stock balance (only has rows where transactions exist)
    query = supabase.table("stock_balance").select("*")
    if view_location_id:
        query = query.eq("location_id", view_location_id)
    stock = query.execute()

    # Build a quick lookup: (location_id, item_id) -> on_hand_qty
    balance_map = {}
    for row in (stock.data or []):
        key = (row.get("location_id"), row.get("item_id"))
        balance_map[key] = float(row.get("on_hand_qty") or 0)

    # Only include items that have stock SOMEWHERE (not every item in the DB)
    active_item_ids = list({row["item_id"] for row in (stock.data or []) if row.get("item_id")})

    # Fetch ALL locations so warehouse always appears
    all_locations = supabase.table("locations").select("id, name, type").order("type").order("name").execute()
    loc_list = all_locations.data or []
    if view_location_id:
        loc_list = [l for l in loc_list if l["id"] == view_location_id]

    # Fetch only the active items
    item_map = {}
    if active_item_ids:
        items = supabase.table("items").select("id, name, sku, conversion_factor").in_("id", active_item_ids).order("name").execute()
        item_list = items.data or []
        item_map = {r["id"]: r for r in item_list}
    else:
        item_list = []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Balance"

    ws.append(["Location", "Type", "Item", "SKU", "On Hand (bags)", "On Hand (kg)", "Status"])
    _style_header_row(ws)

    for loc in loc_list:
        for item in item_list:
            on_hand_kg = balance_map.get((loc["id"], item["id"]), 0.0)
            conversion = float(item.get("conversion_factor") or 1)
            on_hand_bags = int(on_hand_kg / conversion) if conversion > 0 else 0
            location_type = (loc.get("type") or "").capitalize()

            if on_hand_bags <= 0:
                status = "Out of Stock"
            elif on_hand_bags < 5:
                status = "Low"
            else:
                status = "In Stock"

            ws.append([
                loc["name"], location_type, item["name"],
                item.get("sku", ""), on_hand_bags, round(on_hand_kg, 2), status,
            ])

    _auto_column_widths(ws)

    filename = f"stock_balance_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return _workbook_to_response(wb, filename)


# ---------- Transactions Export ----------

@router.get("/transactions/excel")
async def export_transactions_excel(
    days: int = Query(30, le=365),
    type_filter: Optional[str] = None,
    view_location_id: Optional[str] = None,
    user_data: dict = Depends(require_manager),
):
    """Export transaction history as Excel."""
    supabase = get_supabase_admin_client()
    profile = user_data.get("profile", {})
    location_id = get_view_location_id(profile, view_location_id)

    start_date = (datetime.utcnow() - timedelta(days=days)).isoformat()

    query = supabase.table("stock_transactions").select(
        "created_at, type, item_id, qty, unit, notes, items(name, conversion_factor)"
    ).gte("created_at", start_date).order("created_at", desc=True).limit(2000)

    if location_id:
        # Match transactions where this location is either source or destination
        query = query.or_(f"location_id_from.eq.{location_id},location_id_to.eq.{location_id}")
    if type_filter and type_filter != "all":
        query = query.eq("type", type_filter)

    transactions = query.execute()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ws.append(["Date/Time", "Type", "Item", "Qty (bags)", "Qty (kg)", "Notes"])
    _style_header_row(ws)

    for tx in (transactions.data or []):
        dt = tx.get("created_at", "")
        try:
            dt = datetime.fromisoformat(dt.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M")
        except (ValueError, AttributeError):
            pass
        tx_type = (tx.get("type") or "").upper()
        item_info = tx.get("items") or {}
        item_name = item_info.get("name", "")
        qty_kg = float(tx.get("qty") or 0)
        conversion = float(item_info.get("conversion_factor") or 1)
        qty_bags = int(qty_kg / conversion) if conversion > 0 else 0
        notes = tx.get("notes") or ""

        ws.append([dt, tx_type, item_name, qty_bags, round(qty_kg, 2), notes])

    _auto_column_widths(ws)

    filename = f"transactions_{days}d_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return _workbook_to_response(wb, filename)


# ---------- Batches Export ----------

@router.get("/batches/excel")
async def export_batches_excel(
    view_location_id: Optional[str] = None,
    user_data: dict = Depends(require_manager),
):
    """Export active batches as Excel."""
    supabase = get_supabase_admin_client()
    profile = user_data.get("profile", {})
    location_id = get_view_location_id(profile, view_location_id)

    query = supabase.table("stock_batches").select(
        "initial_qty, remaining_qty, received_at, expiry_date, quality_score, status, "
        "items(name, conversion_factor), locations(name), suppliers(name)"
    ).gt("remaining_qty", 0).order("received_at", desc=False)

    if location_id:
        query = query.eq("location_id", location_id)

    batches = query.execute()

    quality_labels = {1: "Good", 2: "Acceptable", 3: "Poor"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batches"

    ws.append([
        "Item", "Location", "Supplier", "Initial (bags)", "Remaining (bags)",
        "Initial (kg)", "Remaining (kg)", "Received", "Expiry", "Quality", "Status",
    ])
    _style_header_row(ws)

    for b in (batches.data or []):
        item_info = b.get("items") or {}
        item_name = item_info.get("name", "")
        conversion = float(item_info.get("conversion_factor") or 1)
        location_name = (b.get("locations") or {}).get("name", "")
        supplier_name = (b.get("suppliers") or {}).get("name", "")
        initial_kg = float(b.get("initial_qty") or 0)
        remaining_kg = float(b.get("remaining_qty") or 0)
        initial_bags = int(initial_kg / conversion) if conversion > 0 else 0
        remaining_bags = int(remaining_kg / conversion) if conversion > 0 else 0
        received = (b.get("received_at") or "")[:10]
        expiry = b.get("expiry_date") or ""
        quality = quality_labels.get(b.get("quality_score"), "Unknown")
        status = (b.get("status") or "").capitalize()

        ws.append([
            item_name, location_name, supplier_name,
            initial_bags, remaining_bags,
            round(initial_kg, 2), round(remaining_kg, 2),
            received, expiry, quality, status,
        ])

    _auto_column_widths(ws)

    filename = f"batches_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return _workbook_to_response(wb, filename)


# ---------- Stock Take Excel Export ----------

@router.get("/stock-take/{stock_take_id}/excel")
async def export_stock_take_excel(
    stock_take_id: str,
    user_data: dict = Depends(require_manager),
):
    """Export stock take results as Excel."""
    supabase = get_supabase_admin_client()

    header = supabase.table("stock_takes").select(
        "*, locations(name)"
    ).eq("id", stock_take_id).single().execute()

    if not header.data:
        raise HTTPException(status_code=404, detail="Stock take not found")

    lines = supabase.table("stock_take_lines").select(
        "*, items(name, conversion_factor)"
    ).eq("stock_take_id", stock_take_id).execute()

    location_name = (header.data.get("locations") or {}).get("name", "Unknown")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Take"

    # Summary rows
    ws.append(["Stock Take Summary"])
    ws.append(["Location", location_name])
    ws.append(["Status", header.data["status"].replace("_", " ").title()])
    ws.append(["Started", header.data["started_at"][:16].replace("T", " ")])
    if header.data.get("completed_at"):
        ws.append(["Completed", header.data["completed_at"][:16].replace("T", " ")])
    ws.append([])

    # Data table
    ws.append(["Item", "Expected (bags)", "Counted (bags)", "Variance (bags)", "Variance %", "Notes"])
    data_header_row = ws.max_row
    for cell in ws[data_header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    for line in (lines.data or []):
        item_info = line.get("items") or {}
        item_name = item_info.get("name", "")
        conversion = float(item_info.get("conversion_factor") or 1)
        expected_kg = float(line.get("expected_qty") or 0)
        expected_bags = int(expected_kg / conversion) if conversion > 0 else 0
        counted = line.get("counted_qty")
        counted_bags = int(float(counted) / conversion) if counted is not None and conversion > 0 else None
        variance = line.get("variance")
        variance_bags = int(float(variance) / conversion) if variance is not None and conversion > 0 else None
        variance_pct = line.get("variance_pct")
        notes = line.get("notes") or ""

        ws.append([
            item_name,
            expected_bags,
            counted_bags if counted_bags is not None else "Not counted",
            variance_bags if variance_bags is not None else "-",
            f"{variance_pct}%" if variance_pct is not None else "-",
            notes,
        ])

    _auto_column_widths(ws)

    filename = f"stock_take_{stock_take_id[:8]}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return _workbook_to_response(wb, filename)


# ---------- Stock Take PDF Export ----------

@router.get("/stock-take/{stock_take_id}/pdf")
async def export_stock_take_pdf(
    stock_take_id: str,
    user_data: dict = Depends(require_manager),
):
    """Export stock take summary as printable PDF."""
    supabase = get_supabase_admin_client()

    header = supabase.table("stock_takes").select(
        "*, locations(name)"
    ).eq("id", stock_take_id).single().execute()

    if not header.data:
        raise HTTPException(status_code=404, detail="Stock take not found")

    # Get initiator name
    initiator_name = "Unknown"
    if header.data.get("initiated_by"):
        profile = supabase.table("profiles").select("full_name").eq(
            "user_id", header.data["initiated_by"]
        ).single().execute()
        if profile.data:
            initiator_name = profile.data.get("full_name", "Unknown")

    lines = supabase.table("stock_take_lines").select(
        "*, items(name, conversion_factor)"
    ).eq("stock_take_id", stock_take_id).execute()

    location_name = (header.data.get("locations") or {}).get("name", "Unknown")
    started = header.data["started_at"][:16].replace("T", " ")
    status = header.data["status"].replace("_", " ").title()

    pdf = FPDF()
    pdf.add_page()

    # Title
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 12, "Stock Take Report", ln=True, align="C")
    pdf.ln(4)

    # Details
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"Location: {location_name}", ln=True)
    pdf.cell(0, 8, f"Initiated by: {initiator_name}", ln=True)
    pdf.cell(0, 8, f"Started: {started}", ln=True)
    pdf.cell(0, 8, f"Status: {status}", ln=True)

    if header.data.get("completed_at"):
        completed = header.data["completed_at"][:16].replace("T", " ")
        pdf.cell(0, 8, f"Completed: {completed}", ln=True)

    # Summary stats
    total_lines = header.data.get("total_lines", 0)
    lines_counted = header.data.get("lines_counted", 0)
    variance_count = header.data.get("variance_count", 0)

    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, f"Items: {total_lines}  |  Counted: {lines_counted}  |  Variances: {variance_count}", ln=True)

    # Table
    pdf.ln(6)

    # Column widths
    col_w = [55, 30, 30, 30, 45]

    # Header
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(79, 70, 229)  # Indigo
    pdf.set_text_color(255, 255, 255)
    headers = ["Item", "Expected (bags)", "Counted (bags)", "Variance (bags)", "Notes"]
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 10, h, border=1, fill=True, align="C")
    pdf.ln()

    # Data rows
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(0, 0, 0)

    for line in (lines.data or []):
        item_info = line.get("items") or {}
        item_name = (item_info.get("name", ""))[:25]
        conversion = float(item_info.get("conversion_factor") or 1)
        expected_kg = float(line.get("expected_qty") or 0)
        expected = f"{int(expected_kg / conversion) if conversion > 0 else 0}"
        counted = f"{int(float(line['counted_qty']) / conversion) if conversion > 0 else 0}" if line.get("counted_qty") is not None else "-"
        variance_val = float(line['variance']) if line.get("variance") is not None else None
        variance = f"{int(variance_val / conversion) if conversion > 0 else 0}" if variance_val is not None else "-"
        notes = (line.get("notes") or "")[:20]

        # Highlight variances
        has_variance = (
            line.get("variance") is not None and abs(float(line["variance"])) > 0.01
        )
        if has_variance:
            pdf.set_fill_color(254, 252, 232)  # Light amber
            fill = True
        else:
            fill = False

        pdf.cell(col_w[0], 8, item_name, border=1, fill=fill)
        pdf.cell(col_w[1], 8, expected, border=1, align="R", fill=fill)
        pdf.cell(col_w[2], 8, counted, border=1, align="R", fill=fill)
        pdf.cell(col_w[3], 8, variance, border=1, align="R", fill=fill)
        pdf.cell(col_w[4], 8, notes, border=1, fill=fill)
        pdf.ln()

    # Signature lines
    pdf.ln(16)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(90, 8, "Counted by: _________________________", ln=False)
    pdf.cell(90, 8, "Approved by: _________________________", ln=True)
    pdf.ln(4)
    pdf.cell(90, 8, "Date: _________________________", ln=False)
    pdf.cell(90, 8, "Date: _________________________", ln=True)

    # Output
    pdf_bytes = pdf.output()
    filename = f"stock_take_{stock_take_id[:8]}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"

    return Response(
        content=bytes(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------- Daily Report PDF ----------

@router.get("/daily-report/pdf")
async def export_daily_report_pdf(
    period_days: int = Query(7, le=90),
    view_location_id: Optional[str] = None,
    user_data: dict = Depends(require_manager),
):
    """Export daily summary report as PDF."""
    supabase = get_supabase_admin_client()
    profile = user_data.get("profile", {})
    location_id = get_view_location_id(profile, view_location_id)

    start_date = (datetime.utcnow() - timedelta(days=period_days)).date()

    # Gather daily data
    daily_data = []
    totals = {"received": 0.0, "issued": 0.0, "wasted": 0.0}

    for i in range(period_days):
        date = start_date + timedelta(days=i)
        date_str = date.isoformat()
        day_start = f"{date_str}T00:00:00"
        day_end = f"{date_str}T23:59:59"

        tx_query = supabase.table("stock_transactions").select(
            "type, qty"
        ).gte("created_at", day_start).lte("created_at", day_end)

        if location_id:
            tx_query = tx_query.or_(
                f"location_id_from.eq.{location_id},location_id_to.eq.{location_id}"
            )

        tx_data = tx_query.execute()

        received = sum(t["qty"] for t in (tx_data.data or []) if t["type"] == "receive")
        issued = sum(t["qty"] for t in (tx_data.data or []) if t["type"] == "issue")
        wasted = sum(t["qty"] for t in (tx_data.data or []) if t["type"] == "waste")

        daily_data.append({
            "date": date_str,
            "received": received,
            "issued": issued,
            "wasted": wasted,
            "net": received - issued - wasted,
        })
        totals["received"] += received
        totals["issued"] += issued
        totals["wasted"] += wasted

    pdf = FPDF()
    pdf.add_page()

    # Title
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 12, f"Daily Report ({period_days} days)", ln=True, align="C")
    pdf.ln(4)

    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"Period: {start_date} to {(datetime.utcnow()).date()}", ln=True)
    pdf.cell(0, 8, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}", ln=True)

    # Summary
    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 11)
    net = totals["received"] - totals["issued"] - totals["wasted"]
    pdf.cell(0, 8, f"Total Received: {totals['received']:.0f} kg  |  Issued: {totals['issued']:.0f} kg  |  Wasted: {totals['wasted']:.0f} kg  |  Net: {net:.0f} kg", ln=True)

    # Table
    pdf.ln(6)
    col_w = [35, 30, 30, 30, 30]

    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(79, 70, 229)
    pdf.set_text_color(255, 255, 255)
    for i, h in enumerate(["Date", "Received", "Issued", "Wasted", "Net Change"]):
        pdf.cell(col_w[i], 10, h, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(0, 0, 0)

    for row in daily_data:
        pdf.cell(col_w[0], 8, row["date"], border=1)
        pdf.cell(col_w[1], 8, f"{row['received']:.0f}", border=1, align="R")
        pdf.cell(col_w[2], 8, f"{row['issued']:.0f}", border=1, align="R")
        pdf.cell(col_w[3], 8, f"{row['wasted']:.0f}", border=1, align="R")
        pdf.cell(col_w[4], 8, f"{row['net']:.0f}", border=1, align="R")
        pdf.ln()

    pdf_bytes = pdf.output()
    filename = f"daily_report_{period_days}d_{datetime.utcnow().strftime('%Y%m%d')}.pdf"

    return Response(
        content=bytes(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
